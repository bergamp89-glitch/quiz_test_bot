import os
import sqlite3
from openpyxl import load_workbook

from database import DB_NAME, init_db, SUBJECT_NAMES


def normalize_subject(subject: str):
    """
    Excel ichidagi fan nomini bot tushunadigan subject_key ga aylantiradi.
    Masalan:
    Iqtisodiyot -> iqtisodiyot
    Mikro-Makro iqtisodiyot -> mikro_makro
    """
    subject = str(subject).strip().lower()

    aliases = {
    "iqtisodiyot": "iqtisodiyot",
    "iqtisodiyot fanidan": "iqtisodiyot",
    "economics": "iqtisodiyot",

    "dm": "dm",
    "dm fanidan": "dm",
    "dm fanidan testlar": "dm",

    "mikro-makro iqtisodiyot": "mikro_makro",
    "mikro makro iqtisodiyot": "mikro_makro",
    "mikro_makro": "mikro_makro",
    "mikmak": "mikro_makro",
    "mikmak fanidan": "mikro_makro",
    "mikroiqtisodiyot": "mikro_makro",
    "makroiqtisodiyot": "mikro_makro",

    "moliya": "moliya",
    "moliya fanidan": "moliya",
    "moliya fanidan testlar": "moliya",

    "statistika": "statistika",
    "statistika fanidan": "statistika",
    "statistika fanidan testlar": "statistika",

    "innovatsion iqtisodiyot": "innovatsion_iqtisodiyot",
    "innovatsion iqtisodiyot fani": "innovatsion_iqtisodiyot",
    "innovatsion_iqtisodiyot": "innovatsion_iqtisodiyot",

    "menejment-marketing": "menejment_marketing",
    "menejment marketing": "menejment_marketing",
    "menejment_marketing": "menejment_marketing",
    "menejment va marketing": "menejment_marketing",
    "marketing menejment": "menejment_marketing",
}

    return aliases.get(subject)


def import_questions_from_excel(filename="questions.xlsx"):
    """
    questions.xlsx faylidagi testlarni quiz.db bazaga yuklaydi.
    Excel ustunlari:
    subject | question | option_a | option_b | option_c | option_d | correct_answer | explanation
    """

    if not os.path.exists(filename):
        print(f"❌ {filename} topilmadi.")
        print("questions.xlsx faylini loyiha papkasiga joylang.")
        return

    # Bazani yaratib olamiz
    init_db()

    workbook = load_workbook(filename)
    sheet = workbook.active

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    # Eski savollarni o‘chirib, Excel savollarni qayta yuklaymiz
    cursor.execute("DELETE FROM questions")

    imported_count = 0
    skipped_count = 0

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            subject, question, option_a, option_b, option_c, option_d, correct_answer, explanation = row
        except ValueError:
            skipped_count += 1
            print(f"⚠️ {row_number}-qator o‘tkazildi: ustunlar soni noto‘g‘ri.")
            continue

        subject_key = normalize_subject(subject)

        if not subject_key:
            skipped_count += 1
            print(f"⚠️ {row_number}-qator o‘tkazildi: fan nomi noto‘g‘ri -> {subject}")
            continue

        if not question or not option_a or not option_b or not option_c or not option_d or not correct_answer:
            skipped_count += 1
            print(f"⚠️ {row_number}-qator o‘tkazildi: bo‘sh katak bor.")
            continue

        correct_answer = str(correct_answer).strip().upper()

        if correct_answer not in ["A", "B", "C", "D"]:
            skipped_count += 1
            print(f"⚠️ {row_number}-qator o‘tkazildi: correct_answer A/B/C/D emas.")
            continue

        subject_name = SUBJECT_NAMES.get(subject_key)

        if not subject_name:
            skipped_count += 1
            print(f"⚠️ {row_number}-qator o‘tkazildi: SUBJECT_NAMES ichida fan yo‘q.")
            continue

        cursor.execute("""
            INSERT INTO questions (
                subject_key,
                subject_name,
                question,
                option_a,
                option_b,
                option_c,
                option_d,
                correct_answer,
                explanation
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            subject_key,
            subject_name,
            str(question).strip(),
            str(option_a).strip(),
            str(option_b).strip(),
            str(option_c).strip(),
            str(option_d).strip(),
            correct_answer,
            str(explanation).strip() if explanation else ""
        ))

        imported_count += 1

    conn.commit()
    conn.close()

    print("✅ Import tugadi.")
    print(f"✅ Yuklangan savollar: {imported_count} ta")
    print(f"⚠️ O‘tkazib yuborilgan qatorlar: {skipped_count} ta")


if __name__ == "__main__":
    import_questions_from_excel()